"""
Excel cue sheet parser and validator.
Pure module — no FastAPI/telnet imports. No side effects.
"""

import io
import re
import tempfile
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font

CUE_NUMBER_RE = re.compile(r"^\d+(\.\d+)?$")

TEMPLATE_COLUMNS = [
    "cue", "label", "fixtures", "intensity",
    "color_r", "color_g", "color_b",
    "pan", "tilt", "focus", "strobe",
    "effect_slot", "effect_value", "effect_rate",
]

HEADER_ALIASES = {
    "큐번호": "cue", "큐": "cue", "cue": "cue",
    "이름": "label", "레이블": "label", "label": "label",
    "조명": "fixtures", "픽스처": "fixtures", "fixtures": "fixtures",
    "밝기": "intensity", "intensity": "intensity",
    "r": "color_r", "color_r": "color_r",
    "g": "color_g", "color_g": "color_g",
    "b": "color_b", "color_b": "color_b",
    "팬": "pan", "pan": "pan",
    "틸트": "tilt", "tilt": "tilt",
    "포커스": "focus", "줌": "focus", "focus": "focus",
    "스트로브": "strobe", "strobe": "strobe",
    "이펙트슬롯": "effect_slot", "effect_slot": "effect_slot",
    "이펙트값": "effect_value", "effect_value": "effect_value",
    "이펙트속도": "effect_rate", "effect_rate": "effect_rate",
}


class RowValidationError(Exception):
    def __init__(self, column: str, message: str):
        self.column = column
        self.message = message
        super().__init__(message)


def _normalize_header(raw: str) -> Optional[str]:
    return HEADER_ALIASES.get(str(raw).strip().lower())


def _parse_cue_number(value) -> str:
    if value is None or str(value).strip() == "":
        raise RowValidationError("cue", "큐번호가 비어있습니다.")
    if isinstance(value, float):
        value = str(int(value)) if value == int(value) else str(value)
    else:
        value = str(value).strip()
    if re.match(r"^\d+\.0$", value):
        value = value[:-2]
    if not CUE_NUMBER_RE.match(value):
        raise RowValidationError("cue", f"큐번호 형식 오류: '{value}' (예: 1, 1.5, 220)")
    return value


def _parse_fixtures(raw) -> list:
    if raw is None or str(raw).strip() == "":
        raise RowValidationError("fixtures", "조명 번호가 비어있습니다.")
    raw = str(raw).strip()
    result = set()
    for token in re.split(r"[,\s]+", raw):
        token = token.strip()
        if not token:
            continue
        if re.match(r"^\d+$", token):
            result.add(int(token))
        elif re.match(r"^\d+-\d+$", token):
            a, b = token.split("-")
            a, b = int(a), int(b)
            if a > b:
                raise RowValidationError("fixtures", f"범위 오류: {token} (시작이 끝보다 큼)")
            result.update(range(a, b + 1))
        else:
            raise RowValidationError("fixtures", f"조명 형식 오류: '{token}' (예: 1,2,5-7)")
    if not result:
        raise RowValidationError("fixtures", "유효한 조명 번호가 없습니다.")
    return sorted(result)


def _parse_int_range(value, field: str, lo: int = 0, hi: int = 100) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    try:
        v = int(float(str(value)))
    except (ValueError, TypeError):
        raise RowValidationError(field, f"{field} 값이 숫자가 아닙니다: '{value}'")
    if not (lo <= v <= hi):
        raise RowValidationError(field, f"{field} 범위 오류({lo}-{hi}): {v}")
    return v


def _validate_row(col_map: dict, row_values: dict, row_index: int) -> dict:
    def get(key):
        idx = col_map.get(key)
        return row_values.get(idx) if idx is not None else None

    cue = _parse_cue_number(get("cue"))
    label = str(get("label") or "").strip()[:40]
    fixtures = _parse_fixtures(get("fixtures"))
    intensity = _parse_int_range(get("intensity"), "intensity")
    pan = _parse_int_range(get("pan"), "pan")
    tilt = _parse_int_range(get("tilt"), "tilt")
    focus = _parse_int_range(get("focus"), "focus")
    strobe = _parse_int_range(get("strobe"), "strobe")
    effect_slot = _parse_int_range(get("effect_slot"), "effect_slot", 1, 99)
    effect_value = _parse_int_range(get("effect_value"), "effect_value")
    effect_rate = _parse_int_range(get("effect_rate"), "effect_rate")

    r = _parse_int_range(get("color_r"), "color_r", 0, 255)
    g = _parse_int_range(get("color_g"), "color_g", 0, 255)
    b = _parse_int_range(get("color_b"), "color_b", 0, 255)
    color_parts = [r, g, b]
    if any(v is not None for v in color_parts):
        if any(v is None for v in color_parts):
            raise RowValidationError("color_r/g/b", "color_r, color_g, color_b는 모두 함께 제공해야 합니다.")
        # 0-255 RGB 입력이면 자동으로 0-100 스케일로 변환
        if any(v > 100 for v in [r, g, b] if v is not None):
            r = round(r / 255 * 100)
            g = round(g / 255 * 100)
            b = round(b / 255 * 100)
        color = {"r": r, "g": g, "b": b}
    else:
        color = None

    effect = None
    if effect_slot is not None and effect_value is not None:
        effect = {"mode": "slot", "slot": effect_slot, "value": effect_value, "rate": effect_rate}
    elif strobe is not None:
        effect = {"mode": "strobe", "strobe": strobe}

    return {
        "row_index": row_index,
        "cue": cue,
        "label": label,
        "fixtures": fixtures,
        "intensity": intensity,
        "color": color,
        "pan": pan,
        "tilt": tilt,
        "focus": focus,
        "effect": effect,
    }


def is_template_format(file_bytes: bytes) -> bool:
    """첫 번째 비어있지 않은 행에 'cue'와 'fixtures' 컬럼이 모두 있으면 템플릿 포맷으로 판단."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        first_nonempty = None
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                first_nonempty = row
                break
        wb.close()
    except Exception:
        return False
    if not first_nonempty:
        return False
    normalized = [_normalize_header(str(c)) for c in first_nonempty if c is not None]
    return "cue" in normalized and "fixtures" in normalized


def parse_workbook(file_bytes: bytes) -> tuple:
    """Returns (rows, errors). rows = validated, errors = [{row_index, column, message}]."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"xlsx 파일을 읽을 수 없습니다: {e}")

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ValueError("파일이 비어있습니다.")

    header_row = all_rows[0]
    col_map: dict = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        norm = _normalize_header(str(cell))
        if norm and norm not in col_map:
            col_map[norm] = i

    if "cue" not in col_map:
        raise ValueError("'cue' 또는 '큐번호' 헤더 컬럼이 없습니다.")
    if "fixtures" not in col_map:
        raise ValueError("'fixtures' 또는 '조명' 헤더 컬럼이 없습니다.")

    rows: list = []
    errors: list = []

    for row_i, row in enumerate(all_rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        row_values = {i: v for i, v in enumerate(row)}
        try:
            parsed = _validate_row(col_map, row_values, row_i)
            rows.append(parsed)
        except RowValidationError as e:
            errors.append({"row_index": row_i, "column": e.column, "message": e.message})

    return rows, errors


def build_commands(row: dict) -> list:
    """검증된 행을 MA2 Telnet 명령어 리스트로 변환."""
    try:
        from .ma2_commands import (
            cmd_select_fixtures, cmd_intensity, cmd_color_rgb,
            cmd_pan, cmd_tilt, cmd_focus, cmd_strobe,
            cmd_effect_slot, cmd_effect_rate, cmd_effect_off,
            cmd_store_cue, cmd_clear_all,
        )
    except ImportError:
        from ma2_commands import (
            cmd_select_fixtures, cmd_intensity, cmd_color_rgb,
            cmd_pan, cmd_tilt, cmd_focus, cmd_strobe,
            cmd_effect_slot, cmd_effect_rate, cmd_effect_off,
            cmd_store_cue, cmd_clear_all,
        )
    cmds: list = []
    cmds.append(cmd_clear_all())
    cmds.append(cmd_select_fixtures(row["fixtures"]))
    if row.get("intensity") is not None:
        cmds.append(cmd_intensity(row["intensity"]))
    if row.get("color"):
        c = row["color"]
        cmds.extend(cmd_color_rgb(c["r"], c["g"], c["b"]))
    if row.get("pan") is not None:
        cmds.append(cmd_pan(row["pan"]))
    if row.get("tilt") is not None:
        cmds.append(cmd_tilt(row["tilt"]))
    if row.get("focus") is not None:
        cmds.append(cmd_focus(row["focus"]))
    eff = row.get("effect")
    if eff:
        if eff["mode"] == "strobe":
            cmds.append(cmd_strobe(eff["strobe"]))
        elif eff["mode"] == "slot":
            cmds.append(cmd_effect_slot(eff["slot"], eff["value"]))
            if eff.get("rate") is not None:
                cmds.append(cmd_effect_rate(eff["rate"]))
    else:
        cmds.append(cmd_effect_off())
    cmds.append(cmd_store_cue(row["cue"]))
    return cmds


def extract_sheet_text(file_bytes: bytes, max_rows: int = 50, max_cols: int = 30) -> str:
    """엑셀 시트를 AI에게 전달할 텍스트 표현으로 변환. 빈 행 제거, 최대 50행×30열."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        lines = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_vals = list(row)[:max_cols]
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue
            cells = [str(v) if v is not None else "" for v in row_vals]
            lines.append(f"행{i+1}: [{', '.join(cells)}]")
            if len(lines) >= max_rows:
                break
    finally:
        wb.close()
    return "\n".join(lines)


def normalize_ai_rows(ai_rows: list) -> list:
    """AI가 반환한 큐 목록을 검증하고 큐 번호를 정리."""
    result = []
    seen_cues = set()
    for idx, row in enumerate(ai_rows, start=1):
        cue = str(row.get("cue") or idx).strip()
        # 소수점 .0 제거
        if re.match(r"^\d+\.0$", cue):
            cue = cue[:-2]
        # 유효하지 않은 큐 번호면 순서 기반 부여
        if not CUE_NUMBER_RE.match(cue):
            cue = str(idx)
        # 중복 큐 번호 처리: 정수 부분에 suffix를 붙여 유효한 번호 생성
        if cue in seen_cues:
            base = cue.split(".")[0]
            suffix = 1
            cue = f"{base}.{suffix}"
            while cue in seen_cues:
                suffix += 1
                cue = f"{base}.{suffix}"
        seen_cues.add(cue)
        result.append({**row, "cue": cue})
    return result


def build_commands_from_ai(row: dict) -> list:
    """AI 파싱 결과(intensity_per_fixture 포함 가능)를 MA2 명령어 리스트로 변환."""
    try:
        from .ma2_commands import (
            cmd_select_fixtures, cmd_intensity, cmd_color_rgb,
            cmd_pan, cmd_tilt, cmd_focus, cmd_store_cue, cmd_clear_all,
        )
    except ImportError:
        from ma2_commands import (
            cmd_select_fixtures, cmd_intensity, cmd_color_rgb,
            cmd_pan, cmd_tilt, cmd_focus, cmd_store_cue, cmd_clear_all,
        )

    cmds = [cmd_clear_all()]
    per_fixture: dict = row.get("intensity_per_fixture") or {}

    if per_fixture:
        # 같은 밝기값을 가진 fixture들을 그룹핑하여 명령어 생성
        groups: dict = {}
        for fixture_str, intensity_val in per_fixture.items():
            try:
                fixture_num = int(fixture_str)
                if intensity_val is None or intensity_val == "":
                    continue
                intensity_int = int(float(intensity_val))
            except (ValueError, TypeError):
                continue
            groups.setdefault(intensity_int, []).append(fixture_num)
        for intensity_val, fixtures in sorted(groups.items(), reverse=True):
            cmds.append(cmd_select_fixtures(sorted(fixtures)))
            cmds.append(cmd_intensity(intensity_val))
    else:
        fixtures = row.get("fixtures") or []
        if fixtures:
            cmds.append(cmd_select_fixtures(fixtures))
        if row.get("intensity") is not None:
            cmds.append(cmd_intensity(int(row["intensity"])))

    color = row.get("color")
    if isinstance(color, dict) and all(k in color for k in ("r", "g", "b")):
        cmds.extend(cmd_color_rgb(int(color["r"]), int(color["g"]), int(color["b"])))

    if row.get("pan") is not None:
        cmds.append(cmd_pan(int(row["pan"])))
    if row.get("tilt") is not None:
        cmds.append(cmd_tilt(int(row["tilt"])))
    if row.get("focus") is not None:
        cmds.append(cmd_focus(int(row["focus"])))

    cmds.append(cmd_store_cue(row["cue"]))
    return cmds


def detect_missing_fields(rows: list) -> list:
    """
    누락 필드를 감지하여 질문이 필요한 큐 목록을 반환한다.
    필수 필드: fixtures/intensity_per_fixture, intensity (fixtures만 있고 intensity 없는 경우)
    label은 전체 50% 이상 누락 시 일괄 이슈로 추가.
    반환: [{"cue": "3", "label": "블루워시", "missing": ["fixtures"]}, ...]
    """
    issues = []
    label_missing_cues = []

    for row in rows:
        missing = []
        has_per_fixture = bool(row.get("intensity_per_fixture"))
        has_fixtures = bool(row.get("fixtures"))

        if not has_per_fixture and not has_fixtures:
            missing.append("fixtures")
        elif has_fixtures and not has_per_fixture and row.get("intensity") is None:
            missing.append("intensity")

        if not str(row.get("label") or "").strip():
            label_missing_cues.append(row["cue"])

        if missing:
            issues.append({
                "cue": row["cue"],
                "label": row.get("label", ""),
                "missing": missing,
            })

    if len(label_missing_cues) > len(rows) / 2:
        issues.append({
            "cue": "__batch_label__",
            "cue_list": label_missing_cues,
            "missing": ["label"],
        })

    return issues


def apply_patches(rows: list, patches: list) -> list:
    """
    GPT가 생성한 패치를 rows에 적용한다.
    patches: [{"cue": "3", "field": "fixtures", "value": [5, 6, 7]}, ...]
    label 배치 패치: {"cue": "__batch_label__", "field": "label", "value": {"1": "등장", "2": "절정", ...}}
    """
    row_map = {r["cue"]: r for r in rows}

    for patch in patches:
        cue_key = str(patch.get("cue", ""))
        field = patch.get("field", "")
        value = patch.get("value")

        if cue_key == "__batch_label__" and field == "label" and isinstance(value, dict):
            for cue_num, lbl in value.items():
                if cue_num in row_map:
                    row_map[cue_num]["label"] = str(lbl).strip()[:40]
            continue

        if cue_key not in row_map:
            continue

        row = row_map[cue_key]

        if field == "fixtures":
            try:
                parsed = _parse_fixtures(str(value) if not isinstance(value, list) else " ".join(str(v) for v in value))
                row["fixtures"] = parsed
            except RowValidationError:
                pass
        elif field == "intensity":
            try:
                v = int(float(str(value)))
                row["intensity"] = max(0, min(100, v))
            except (ValueError, TypeError):
                pass
        elif field == "label":
            row["label"] = str(value or "").strip()[:40]
        elif field == "color" and isinstance(value, dict):
            try:
                row["color"] = {
                    "r": max(0, min(100, int(value.get("r", 0)))),
                    "g": max(0, min(100, int(value.get("g", 0)))),
                    "b": max(0, min(100, int(value.get("b", 0)))),
                }
            except (ValueError, TypeError):
                pass
        elif field == "fade":
            try:
                row["fade"] = float(value)
            except (ValueError, TypeError):
                pass

    return list(row_map.values())


def generate_template_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cues"
    ws.append(TEMPLATE_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(["1", "오프닝", "1-4", 80, 100, 0, 0, 50, 50, "", "", "", "", ""])
    ws.append(["2", "블루워시", "5,6,7", 60, 0, 0, 100, "", "", "", "", "", "", ""])
    wb.save(path)
